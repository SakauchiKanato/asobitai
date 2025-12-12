import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
import shutil
import os
from datetime import datetime

OUTPUT_DIR = "output"

def create_from_yolo(yolo_results, unit_list, template_path, student_name=None):
    """
    YOLOの生データを受け取って、Excelを作成する便利関数
    """
    # 1. YOLOのデータを「Y座標」順にソート
    sorted_yolo = sorted(yolo_results, key=lambda item: item.get('y', 0))

    # 2. データの整形
    formatted_data = []
    for i, item in enumerate(sorted_yolo):
        unit_name = unit_list[i] if i < len(unit_list) else "未設定"
        formatted_data.append({
            "num": i + 1,
            "unit": unit_name,
            "result": item.get('class', '?')
        })
    
    # 3. レポート作成へ
    return create_report(formatted_data, template_path, student_name)


def create_report(grading_data_list, template_path, student_name=None):
    """
    Excelを作成・編集する関数（テンプレートがなくても自動生成するように修正）
    """
    
    # --- 準備 ---
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"score_report_{timestamp}.xlsx"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    save_path = os.path.join(OUTPUT_DIR, filename)
    

    # ★ここが重要: ファイルがあるかチェックし、なければ新規作成する

    if template_path and os.path.exists(template_path):
        # テンプレートがある場合 -> コピーして使う
        print(f"テンプレートを使用します: {template_path}")
        shutil.copy(template_path, save_path)
        wb = openpyxl.load_workbook(save_path)
    else:
        # ★テンプレートがない場合 -> 白紙から新規作成する！
        print(f"テンプレートが見つからないため、新規作成します: {save_path}")
        wb = openpyxl.Workbook()
        ws = wb.active
        # 最低限のヘッダーを書き込んでおく
        ws.append(["No.", "単元(Unit)", "結果(Result)", "判定(Status)"])
        wb.save(save_path)

    ws = wb.active

    # 名前書き込み 
    if student_name:
        # 書き込む場所（B1）
        ws['B1'] = student_name
        print(f"氏名「{student_name}」を書き込みました")

    # 1. ヘッダー行を探す
    header_row = None
    col_map = {} 

    for r in range(1, 21):
        row_values = [str(ws.cell(row=r, column=c).value or "") for c in range(1, 20)]
        row_text = "".join(row_values)
        
        if "問" in row_text or "No" in row_text or "判定" in row_text or "Status" in row_text:
            header_row = r
            for c in range(1, 20):
                val = str(ws.cell(row=r, column=c).value or "")
                if "問" in val or "No" in val: col_map["num"] = c
                elif "単元" in val or "ジャンル" in val or "Unit" in val: col_map["unit"] = c
                elif "結果" in val or "Result" in val or "マーク" in val: col_map["result"] = c
                elif "判定" in val or "Status" in val or "正解" in val: col_map["status"] = c
            break
    
    if header_row is None:
        header_row = 1
        col_map = {"num": 1, "unit": 2, "result": 3, "status": 4}

    # 2. 書き込み
    current_row = header_row + 1
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    unit_stats = {} 

    for item in grading_data_list:
        q_num = item.get("num", "-")
        unit = item.get("unit", "未設定")
        mark = item.get("result", "?")

        status = "不正解"
        if mark == "◯": status = "正解"
        elif mark == "△": status = "部分点"
        
        if unit not in unit_stats: unit_stats[unit] = [0, 0]
        if mark == "◯": unit_stats[unit][0] += 1
        unit_stats[unit][1] += 1

        # 列マップに従って書き込み
        if "num" in col_map:
            cell = ws.cell(row=current_row, column=col_map["num"])
            cell.value = f"問{q_num}" if isinstance(q_num, int) else q_num
            cell.border = thin_border
        
        if "unit" in col_map:
            cell = ws.cell(row=current_row, column=col_map["unit"])
            cell.value = unit
            cell.border = thin_border

        if "result" in col_map:
            cell = ws.cell(row=current_row, column=col_map["result"])
            cell.value = mark
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")
            if mark == "◯": cell.font = Font(color="FF0000", bold=True)
            elif mark == "❌": cell.font = Font(color="0000FF", bold=True)

        if "status" in col_map:
            cell = ws.cell(row=current_row, column=col_map["status"])
            cell.value = status
            cell.border = thin_border

        current_row += 1

    # 3. グラフ作成
    max_used_col = max(col_map.values()) if col_map else 5
    graph_start_col = max_used_col + 2 

    row_idx = 1
    ws.cell(row=row_idx, column=graph_start_col, value="単元")
    ws.cell(row=row_idx, column=graph_start_col+1, value="正答率")
    row_idx += 1

    for u_name, stats in unit_stats.items():
        rate = (stats[0] / stats[1]) * 100 if stats[1] > 0 else 0
        ws.cell(row=row_idx, column=graph_start_col, value=u_name)
        ws.cell(row=row_idx, column=graph_start_col+1, value=rate)
        row_idx += 1
    
    chart = BarChart()
    chart.title = "単元別正答率"
    data = Reference(ws, min_col=graph_start_col+1, min_row=1, max_row=row_idx-1)
    cats = Reference(ws, min_col=graph_start_col, min_row=2, max_row=row_idx-1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    ws.add_chart(chart, f"{openpyxl.utils.get_column_letter(graph_start_col)}10")

    wb.save(save_path)
    print(f"Excel作成完了: {save_path}")
    return save_path